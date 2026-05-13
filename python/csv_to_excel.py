import os
from openpyxl import Workbook, load_workbook
import csv

import matplotlib.pyplot as plt

import numpy as np

csv_file = "ble_log_csv_copy.csv"
ble_file = "ble_log.xlsx"

BAR_WIDTH = 30
cnt = 0
x = []
y = []
s = []

def count_rows(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        csv_reader = csv.DictReader(f)
        return sum(1 for r in csv_reader)


def print_progress(current, total):
    percent = int((current / total) * 100)

    filled = int(BAR_WIDTH * percent / 100)
    empty = BAR_WIDTH - filled

    bar = "█" * filled + " " * empty
    print("\r[%s] %d%%" % (bar, percent), end="")


def hampel_filter(data, window_size=71, n_sigmas=3):
    x = np.array(data, dtype=float)
    n = len(x)
    result = x.copy()

    for i in range(window_size,n - window_size):
        window = x[i-window_size:i+window_size+1]
        median = np.median(window)
        mad = np.median(np.abs(window-median))

        if mad == 0:
            continue

        threshold = n_sigmas * 1.4826 * mad

        if abs(x[i]-median) > threshold:
            result[i] = median
    return result


total_rows = count_rows(csv_file)
print(f"Всего {total_rows} записей")

print("Начало форматирования")

'''
if not os.path.exists(ble_file):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(["MAC", "RSSI", "Monotonic", "UTC (ESP)",
               "UTC Now (SYS)", "UTC Now (NTP)",
               "Time sync", "UTC NTP - UTC ESP", ])
    wb.save(ble_file)
'''
with open(csv_file, mode='r', newline='') as f:
    # Создаем объект DictReader
    csv_reader = csv.DictReader(f)

    '''
    wb = load_workbook(ble_file)
    ws = wb.active
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 17
    '''

    for row in csv_reader:
        # Теперь row — это словарь
        # Например: {'name': 'John Smith', 'department': 'Accounting', 'birth_month': 'November'}
        cnt += 1
        x.append(cnt)
        y.append(row['UTC NTP - UTC ESP'])
        s.append(float(row['Monotonic'])/1000)
        #print(f"{cnt}------------------\nUTC (ESP): {row['UTC (ESP)']}\nUTC Now (NTP): {row['UTC Now (NTP)']}\nTime Sync: {row['Time sync']}\nUTC NTP - UTC ESP: {row['UTC NTP - UTC ESP']}\n")
        '''
        ws.append([row['MAC'],
                   int(row['RSSI']),
                   int(row['Monotonic']),
                   row['UTC (ESP)'],
                   row['UTC Now (SYS)'],
                   row['UTC Now (NTP)'],
                   int(row['Time sync']),
                   float(row['UTC NTP - UTC ESP'])
                   ])'''
        if cnt % 500 == 0 or cnt == total_rows:
            print_progress(cnt, total_rows)
    #wb.save(ble_file)
    print("\nОтформатироввано в Excel")

    yy=y[28000:29000]
    ss=s[28000:29000]
    #data = np.array(y, dtype=float)
    #x = np.arange(1, len(data) + 1)
    filtered_data = hampel_filter(y)
    x = np.arange(1, len(filtered_data) + 1)
    x_s = np.array(s, dtype=float)

    plt.figure(figsize=(12, 6))

    plt.plot(x_s, filtered_data, linewidth=1)

    plt.xlabel("Monotonic ESP s")
    plt.ylabel("Signal value")
    plt.title("Signal trend")

    plt.grid(True)

    plt.tight_layout()

    plt.show()